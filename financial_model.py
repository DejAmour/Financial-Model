#!/usr/bin/env python3
"""
Joule Financial Model Generator
================================
Generates a fully linked three-statement financial model (Income Statement,
Balance Sheet, Cash Flow Statement) for a green energy infrastructure
financing company.

Usage:
    python financial_model.py

Output:
    financial_model.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

# ==============================================================================
# ALL ASSUMPTIONS DEFINED IN A SINGLE DICTIONARY
# ==============================================================================

ASSUMPTIONS = {
    # --- Model Parameters ---
    "years": [2026, 2027, 2028, 2029, 2030],

    # --- Fee Structure ---
    "upfront_fee_rate": 0.02,           # 2% of funds raised
    "completion_fee_rate": 0.005,       # 0.5% of funds raised, spread over build period
    "project_completion_years": 3,      # 3-year straight-line build period (dynamic)
    "kwh_fee_rate": 0.01,               # £0.01 per kWh
    "kwh_revenue_rate": 0.0212,         # ~2.12% of funds raised (annual, post-completion)
    "kwh_contract_duration": 10,        # 10-year contract; model shows 5 years only

    # --- Base Case Funds Raised (Total = Projects × Avg Raise) ---
    # 2026: 3 × £6M = £18M | 2027: 7 × £8M = £56M | 2028: 15 × £10M = £150M
    # 2029: 24 × £13M = £312M | 2030: 36 × £16M = £576M
    "base_funds_raised": [18_000_000, 56_000_000, 150_000_000, 312_000_000, 576_000_000],

    # --- Scenario Multipliers (applied to Total Funds Raised only) ---
    "worst_case_multiplier": 0.6,
    "best_case_multiplier": 1.4,

    # --- COGS: Smart Contract Deployment & Audits (fixed per year, dynamic) ---
    "smart_contract_costs": [185_000, 170_000, 105_000, 115_000, 130_000],
    "hosting_api_rate": 0.015,          # 1.5% of gross revenue

    # --- Operating Expenses: Year 1 (2026) Hardcoded ---
    "opex_y1_salaries": 225_000,
    "opex_y1_marketing": 50_000,
    "opex_y1_legal": 100_000,
    "opex_y1_insurance": 20_000,
    "opex_y1_rd": 50_000,

    # --- Operating Expenses: Years 2-5 (% of Gross Revenue) ---
    # Order: [2027, 2028, 2029, 2030]
    "opex_salaries_pct":   [0.25, 0.25, 0.25, 0.25],
    "opex_marketing_pct":  [0.12, 0.07, 0.07, 0.07],
    "opex_legal_pct":      [0.08, 0.08, 0.08, 0.08],
    "opex_insurance_pct":  [0.02, 0.01, 0.01, 0.01],
    "opex_rd_pct":         [0.19, 0.11, 0.05, 0.05],

    # --- Tax Assumptions ---
    "small_profits_rate": 0.19,         # 19% for profits ≤ £50,000
    "main_rate": 0.25,                  # 25% for profits > £250,000
    "lower_profits_threshold": 50_000,
    "upper_profits_threshold": 250_000,
    "tax_payment_lag_months": 9,        # Tax paid ~9 months after year-end (next year)
    "loss_carry_forward_limit": 5_000_000,  # £5M annual deduction limit

    # --- Equity Raises ---
    "equity_raises": [500_000, 1_500_000, 4_000_000, 0, 9_000_000],

    # --- Working Capital ---
    "dso_days": 45,                     # Days Sales Outstanding
    "dpo_days": 30,                     # Days Payable Outstanding

    # --- Opening Balances ---
    # Opening cash = £0 before first equity raise; Y1 equity raise provides initial cash.
    "opening_cash": 0,

    # --- Other ---
    "capex": 0,                         # No CapEx (asset-light model)
    "project_attrition_rate": None,     # Placeholder for future use

    # --- Valuation Model Assumptions ---
    "valuation_date": "2025-08-01",
    "exit_year": 5,
    "ev_revenue_multiple": 7,
    "financial_debt_at_exit": 0,
    "required_return_y1_3": 0.50,
    "required_return_y3_6": 0.35,
    "required_return_y6_10": 0.25,
    "revenue_cagr_y5_y10": 0.60,
    "equity_growth_rate_y5_10": 0.60,

    # --- Cap Table Assumptions ---
    "fx_rate_usd_gbp": 0.7577,
    "founder_shares": 1_000_000,
    "funding_rounds": [
        {"name": "Pre-seed T1", "date": "2025-10-01", "ownership_pct": 0.040652},
        {"name": "Pre-seed T2", "date": "2026-03-01", "ownership_pct": 0.029269},
        {"name": "Seed",        "date": "2026-10-01", "ownership_pct": 0.137200},
        {"name": "Series A T1", "date": "2027-10-01", "ownership_pct": 0.095278},
        {"name": "Series A T2", "date": "2028-10-01", "ownership_pct": 0.049624},
        {"name": "Series B",    "date": "2029-10-01", "ownership_pct": 0.129228},
    ],

    # --- kWh Revenue Calculation Basis (Source: UK DESNZ) ---
    "wind_turbine_capacity_mw": 2.5,    # Midpoint of 2-3 MW (UK DESNZ)
    "annual_output_mwh": 7_000,         # Midpoint of 6,000-8,000 MWh per turbine p.a.
    "cost_per_turbine": 3_300_000,      # Midpoint of £2.6M-£4M per turbine
    # Formula: Funds ÷ £3.3M × 7,000 MWh × £0.01/kWh ≈ 2.12% of Funds (annually)
}


# ==============================================================================
# STYLE CONSTANTS (FAST Modelling Standards)
# ==============================================================================

BLUE   = Font(name="Calibri", color="0000FF", size=10)          # Static inputs
BLACK  = Font(name="Calibri", color="000000", size=10)          # Formulas/calculations
GREEN  = Font(name="Calibri", color="008000", size=10)          # Cross-sheet links
BOLD_BLACK = Font(name="Calibri", color="000000", size=10, bold=True)
BOLD_WHITE = Font(name="Calibri", color="FFFFFF", size=10, bold=True)
TITLE_FONT = Font(name="Calibri", color="FFFFFF", size=12, bold=True)

SECTION_FILL   = PatternFill("solid", fgColor="1F4E79")  # Dark blue header
SUBSECT_FILL   = PatternFill("solid", fgColor="BDD7EE")  # Light blue sub-header
TOTAL_FILL     = PatternFill("solid", fgColor="D9E1F2")  # Light fill for totals
CHECK_OK_FILL  = PatternFill("solid", fgColor="C6EFCE")  # Green for balanced
CHECK_ERR_FILL = PatternFill("solid", fgColor="FFC7CE")  # Red for out of balance
WARN_FILL      = PatternFill("solid", fgColor="FFEB9C")  # Yellow for notes

GBP_FORMAT  = '£#,##0'
GBP_FORMAT2 = '£#,##0.00'
PCT_FORMAT  = '0.0%'
NUM_FORMAT  = '#,##0'

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")

THIN = Side(border_style="thin", color="000000")
THIN_BORDER = Border(bottom=THIN)
BOX_BORDER  = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


# ==============================================================================
# HELPER FUNCTIONS
# ==============================================================================

def col(n: int) -> str:
    """Return Excel column letter for 1-based column index."""
    return get_column_letter(n)


def set_cell(ws, row: int, column: int, value, font=None, fill=None,
             alignment=None, number_format=None, border=None):
    """Write a value to a cell and apply optional styles."""
    c = ws.cell(row=row, column=column, value=value)
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if alignment:
        c.alignment = alignment
    if number_format:
        c.number_format = number_format
    if border:
        c.border = border
    return c


def write_section_header(ws, row: int, label: str, num_cols: int = 8):
    """Write a dark-blue section header spanning num_cols columns."""
    c = ws.cell(row=row, column=1, value=label)
    c.font = TITLE_FONT
    c.fill = SECTION_FILL
    c.alignment = LEFT
    for col_idx in range(2, num_cols + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = SECTION_FILL


def write_subheader(ws, row: int, label: str, num_cols: int = 8):
    """Write a light-blue sub-section header."""
    c = ws.cell(row=row, column=1, value=label)
    c.font = Font(name="Calibri", color="1F4E79", size=10, bold=True, italic=True)
    c.fill = SUBSECT_FILL
    for col_idx in range(2, num_cols + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = SUBSECT_FILL


def write_year_headers(ws, row: int, years: list, start_col: int = 3):
    """Write year headers in bold black across columns."""
    for i, yr in enumerate(years):
        c = ws.cell(row=row, column=start_col + i, value=yr)
        c.font = BOLD_BLACK
        c.alignment = CENTER
        c.border = THIN_BORDER


def add_comment(ws, row: int, column: int, text: str, author: str = "Joule Model"):
    """Add an Excel comment to a cell."""
    comment = Comment(text, author)
    comment.width  = 300
    comment.height = 120
    ws.cell(row=row, column=column).comment = comment


# ==============================================================================
# 1. CREATE ASSUMPTIONS SHEET
# ==============================================================================

def create_assumptions_sheet(wb: openpyxl.Workbook, assumptions: dict):
    """
    Creates the Assumptions tab with all model inputs.
    Blue font = static input, clearly labelled sections.
    Contains the scenario dropdown (Worst / Base / Best).
    """
    ws = wb.create_sheet("Assumptions")

    # --- Column widths ---
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 12
    for c in ["C", "D", "E", "F", "G"]:
        ws.column_dimensions[c].width = 16
    ws.column_dimensions["H"].width = 36

    # ── Title ──────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 24
    set_cell(ws, 1, 1, "JOULE FINANCIAL MODEL — ASSUMPTIONS",
             font=TITLE_FONT, fill=SECTION_FILL, alignment=LEFT)
    for c in range(2, 9):
        ws.cell(row=1, column=c).fill = SECTION_FILL

    # ── Scenario Selector ──────────────────────────────────────────────────────
    write_section_header(ws, 3, "SCENARIO SELECTOR")
    set_cell(ws, 4, 1, "Select Scenario:", font=BLACK, alignment=LEFT)
    set_cell(ws, 4, 3, "Base", font=BLUE, alignment=CENTER)
    set_cell(ws, 4, 8,
             "Choose: Worst / Base / Best  (affects Total Funds Raised only)",
             font=BLACK, alignment=LEFT)

    # Data-validation dropdown
    dv = DataValidation(
        type="list",
        formula1='"Worst,Base,Best"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid",
        error="Select Worst, Base, or Best"
    )
    dv.sqref = "C4"
    ws.add_data_validation(dv)

    # ── Fee Structure ──────────────────────────────────────────────────────────
    write_section_header(ws, 6, "FEE STRUCTURE")
    rows_fee = [
        ("Upfront Fee Rate (% of Funds Raised)",                  assumptions["upfront_fee_rate"],         PCT_FORMAT),
        ("Completion Fee Rate (% of Funds Raised, total)",         assumptions["completion_fee_rate"],       PCT_FORMAT),
        ("Project Completion Time (Years, straight-line spread)",  assumptions["project_completion_years"],  NUM_FORMAT),
        ("kWh Fee Rate (£ per kWh)",                               assumptions["kwh_fee_rate"],              GBP_FORMAT2),
        ("kWh Revenue Rate (% of Funds Raised, annual recurring)", assumptions["kwh_revenue_rate"],          PCT_FORMAT),
        ("kWh Contract Duration (Years, model shows 5 years only)",assumptions["kwh_contract_duration"],     NUM_FORMAT),
    ]
    for i, (label, val, fmt) in enumerate(rows_fee, start=7):
        set_cell(ws, i, 1, label,  font=BLACK, alignment=LEFT)
        set_cell(ws, i, 3, val,    font=BLUE,  alignment=CENTER, number_format=fmt)

    add_comment(ws, 7, 3,
        "Upfront fee: 2% charged on total funds raised in that year.\n"
        "Source: Company fee schedule.")
    add_comment(ws, 8, 3,
        "Completion fee: 0.5% of total project funds raised, recognised\n"
        "evenly (straight-line) over the 3-year project build period.\n"
        "Annual recognition = 0.5% / 3 = 0.1667% per year.")
    add_comment(ws, 11, 3,
        "kWh Revenue Calculation Basis (Source: UK DESNZ):\n"
        "Wind turbine: 2-3 MW capacity, 6,000-8,000 MWh/yr output.\n"
        "Midpoints used: 2.5 MW, 7,000 MWh, £3.3M cost per turbine.\n"
        "Formula: Funds ÷ £3.3M × 7,000 MWh × £0.01/kWh ≈ 2.12% of Funds p.a.\n"
        "See UK DESNZ LCOE estimates: https://www.gov.uk/government/collections/energy-statistics")
    add_comment(ws, 12, 3,
        "kWh contracts are assumed to run for 10 years post-completion.\n"
        "This model forecasts only 5 years; kWh revenue continues beyond the\n"
        "model period and is not captured here.")

    # ── Base Case Funds Raised ─────────────────────────────────────────────────
    write_section_header(ws, 14, "BASE CASE — TOTAL FUNDS RAISED PER YEAR")
    set_cell(ws, 15, 1, "Year →", font=BOLD_BLACK, alignment=LEFT)
    write_year_headers(ws, 15, assumptions["years"])
    set_cell(ws, 16, 1, "Total Funds Raised (Base Case)", font=BLACK, alignment=LEFT)
    for i, val in enumerate(assumptions["base_funds_raised"]):
        set_cell(ws, 16, 3 + i, val, font=BLUE, alignment=RIGHT, number_format=GBP_FORMAT)

    add_comment(ws, 16, 3,
        "Base case: 3 projects × £6M avg raise = £18M\n"
        "2027: 7 × £8M = £56M | 2028: 15 × £10M = £150M\n"
        "2029: 24 × £13M = £312M | 2030: 36 × £16M = £576M")

    # ── Scenario Multipliers ───────────────────────────────────────────────────
    write_section_header(ws, 18, "SCENARIO MULTIPLIERS (applied to Total Funds Raised)")
    rows_scen = [
        ("Worst Case Multiplier", assumptions["worst_case_multiplier"]),
        ("Best Case Multiplier",  assumptions["best_case_multiplier"]),
    ]
    for i, (label, val) in enumerate(rows_scen, start=19):
        set_cell(ws, i, 1, label, font=BLACK, alignment=LEFT)
        set_cell(ws, i, 3, val,   font=BLUE,  alignment=CENTER, number_format="0.0")

    add_comment(ws, 19, 3, "Worst case = Base × 0.6  (40% haircut on Funds Raised)")
    add_comment(ws, 20, 3, "Best case  = Base × 1.4  (40% uplift on Funds Raised)")

    # ── COGS ──────────────────────────────────────────────────────────────────
    write_section_header(ws, 22, "COST OF GOODS SOLD (COGS)")
    set_cell(ws, 23, 1, "Year →", font=BOLD_BLACK, alignment=LEFT)
    write_year_headers(ws, 23, assumptions["years"])
    set_cell(ws, 24, 1, "Smart Contract Deployment & Audits (£)", font=BLACK, alignment=LEFT)
    for i, val in enumerate(assumptions["smart_contract_costs"]):
        set_cell(ws, 24, 3 + i, val, font=BLUE, alignment=RIGHT, number_format=GBP_FORMAT)
    set_cell(ws, 25, 1, "Hosting & API Costs (% of Gross Revenue)", font=BLACK, alignment=LEFT)
    set_cell(ws, 25, 3, assumptions["hosting_api_rate"], font=BLUE, alignment=CENTER, number_format=PCT_FORMAT)
    set_cell(ws, 25, 4, "(same % applies each year)", font=BLACK, alignment=LEFT)

    # ── OpEx Year 1 ────────────────────────────────────────────────────────────
    write_section_header(ws, 27, "OPERATING EXPENSES — YEAR 1 / 2026 (HARDCODED £)")
    rows_opex1 = [
        ("Salaries",           assumptions["opex_y1_salaries"]),
        ("Marketing & Growth", assumptions["opex_y1_marketing"]),
        ("Legal & Regulatory", assumptions["opex_y1_legal"]),
        ("Insurance",          assumptions["opex_y1_insurance"]),
        ("R&D",                assumptions["opex_y1_rd"]),
    ]
    for i, (label, val) in enumerate(rows_opex1, start=28):
        set_cell(ws, i, 1, label, font=BLACK, alignment=LEFT)
        set_cell(ws, i, 3, val,   font=BLUE,  alignment=RIGHT, number_format=GBP_FORMAT)

    # ── OpEx Years 2-5 ─────────────────────────────────────────────────────────
    write_section_header(ws, 34, "OPERATING EXPENSES — YEARS 2–5 (% OF GROSS REVENUE)")
    set_cell(ws, 35, 1, "Year →", font=BOLD_BLACK, alignment=LEFT)
    for i, yr in enumerate(assumptions["years"][1:]):     # 2027–2030
        set_cell(ws, 35, 4 + i, yr, font=BOLD_BLACK, alignment=CENTER, border=THIN_BORDER)
    rows_opex_pct = [
        ("Salaries %",           assumptions["opex_salaries_pct"]),
        ("Marketing & Growth %", assumptions["opex_marketing_pct"]),
        ("Legal & Regulatory %", assumptions["opex_legal_pct"]),
        ("Insurance %",          assumptions["opex_insurance_pct"]),
        ("R&D %",                assumptions["opex_rd_pct"]),
    ]
    for i, (label, pcts) in enumerate(rows_opex_pct, start=36):
        set_cell(ws, i, 1, label, font=BLACK, alignment=LEFT)
        for j, pct in enumerate(pcts):
            set_cell(ws, i, 4 + j, pct, font=BLUE, alignment=CENTER, number_format=PCT_FORMAT)

    # ── Tax Assumptions ────────────────────────────────────────────────────────
    write_section_header(ws, 42, "TAX ASSUMPTIONS")
    rows_tax = [
        ("Small Profits Rate (profits ≤ £50,000)",         assumptions["small_profits_rate"],         PCT_FORMAT),
        ("Main Rate (profits > £250,000)",                  assumptions["main_rate"],                  PCT_FORMAT),
        ("Lower Profits Threshold (£)",                     assumptions["lower_profits_threshold"],    GBP_FORMAT),
        ("Upper Profits Threshold (£)",                     assumptions["upper_profits_threshold"],    GBP_FORMAT),
        ("Tax Payment Lag (months after year-end)",         assumptions["tax_payment_lag_months"],     NUM_FORMAT),
        ("Loss Carry-Forward Annual Deduction Limit (£)",   assumptions["loss_carry_forward_limit"],   GBP_FORMAT),
    ]
    for i, (label, val, fmt) in enumerate(rows_tax, start=43):
        set_cell(ws, i, 1, label, font=BLACK, alignment=LEFT)
        set_cell(ws, i, 3, val,   font=BLUE,  alignment=CENTER, number_format=fmt)

    add_comment(ws, 48, 3,
        "UK HMRC Loss Carry-Forward Rules (post-April 2017):\n"
        "Trading losses may be carried forward indefinitely.\n"
        "From 1 April 2017, losses can only offset up to 50% of annual\n"
        "profits exceeding £5M (the deductions allowance). This model\n"
        "applies the £5M annual deduction limit as a simplification.\n"
        "Source: HMRC CTM04800 / Finance Act 2017.")

    # ── Equity Raises ─────────────────────────────────────────────────────────
    write_section_header(ws, 50, "EQUITY RAISES")
    set_cell(ws, 51, 1, "Year →", font=BOLD_BLACK, alignment=LEFT)
    write_year_headers(ws, 51, assumptions["years"])
    set_cell(ws, 52, 1, "Equity Raised (£)", font=BLACK, alignment=LEFT)
    for i, val in enumerate(assumptions["equity_raises"]):
        set_cell(ws, 52, 3 + i, val, font=BLUE, alignment=RIGHT, number_format=GBP_FORMAT)

    # ── Working Capital ────────────────────────────────────────────────────────
    write_section_header(ws, 54, "WORKING CAPITAL")
    rows_wc = [
        ("Days Sales Outstanding / DSO (days)",  assumptions["dso_days"],    NUM_FORMAT),
        ("Days Payable Outstanding / DPO (days)", assumptions["dpo_days"],   NUM_FORMAT),
    ]
    for i, (label, val, fmt) in enumerate(rows_wc, start=55):
        set_cell(ws, i, 1, label, font=BLACK, alignment=LEFT)
        set_cell(ws, i, 3, val,   font=BLUE,  alignment=CENTER, number_format=fmt)

    # ── Opening Balances ───────────────────────────────────────────────────────
    write_section_header(ws, 58, "OPENING BALANCES")
    set_cell(ws, 59, 1, "Opening Cash (before first equity raise, 2026)", font=BLACK, alignment=LEFT)
    set_cell(ws, 59, 3, assumptions["opening_cash"], font=BLUE, alignment=RIGHT, number_format=GBP_FORMAT)
    set_cell(ws, 59, 4,
             "Note: Y1 equity raise (£500K) provides initial operating cash (shown in Financing)",
             font=BLACK, alignment=LEFT)

    # ── Other Assumptions ─────────────────────────────────────────────────────
    write_section_header(ws, 61, "OTHER ASSUMPTIONS")
    set_cell(ws, 62, 1, "Capital Expenditure / CapEx (£)", font=BLACK, alignment=LEFT)
    set_cell(ws, 62, 3, assumptions["capex"], font=BLUE, alignment=RIGHT, number_format=GBP_FORMAT)
    set_cell(ws, 62, 4, "Asset-light model; no CapEx planned", font=BLACK, alignment=LEFT)
    set_cell(ws, 63, 1, "Project Attrition Rate (%)", font=BLACK, alignment=LEFT)
    set_cell(ws, 63, 3, "", font=BLUE, alignment=CENTER, number_format=PCT_FORMAT,
             fill=WARN_FILL)
    set_cell(ws, 63, 4,
             "Placeholder — leave blank. To be populated in future model iterations.",
             font=BLACK, alignment=LEFT)

    # ── kWh Revenue Calculation Basis ─────────────────────────────────────────
    write_section_header(ws, 65, "kWh REVENUE CALCULATION BASIS  (Source: UK DESNZ)")
    rows_kwh = [
        ("Wind Turbine Capacity (MW, midpoint of 2–3 MW)",        assumptions["wind_turbine_capacity_mw"],  "0.0"),
        ("Annual Output per Turbine (MWh, midpoint of 6,000–8,000)", assumptions["annual_output_mwh"],     NUM_FORMAT),
        ("Average Cost per Turbine (£, midpoint of £2.6M–£4M)",   assumptions["cost_per_turbine"],         GBP_FORMAT),
    ]
    for i, (label, val, fmt) in enumerate(rows_kwh, start=66):
        set_cell(ws, i, 1, label, font=BLACK, alignment=LEFT)
        set_cell(ws, i, 3, val,   font=BLUE,  alignment=CENTER, number_format=fmt)

    set_cell(ws, 69, 1, "kWh Revenue Formula:", font=BOLD_BLACK, alignment=LEFT)
    set_cell(ws, 69, 3,
             "Funds Raised ÷ £3.3M × 7,000 MWh × £0.01/kWh  =  2.12% of Funds Raised p.a.",
             font=BLACK, alignment=LEFT)
    ws.merge_cells("C69:H69")
    set_cell(ws, 70, 1, "Reference:", font=BOLD_BLACK, alignment=LEFT)
    set_cell(ws, 70, 3,
             "UK DESNZ LCOE estimates — https://www.gov.uk/government/collections/energy-statistics",
             font=Font(name="Calibri", color="0563C1", size=10, underline="single"),
             alignment=LEFT)
    ws.merge_cells("C70:H70")
    set_cell(ws, 71, 1,
             "Note: kWh contracts assumed 10 years; model period = 5 years only. "
             "Ongoing revenues not modelled beyond 2030.",
             font=Font(name="Calibri", color="FF0000", size=10, italic=True),
             alignment=LEFT)
    ws.merge_cells("A71:H71")

    # Freeze top row
    ws.freeze_panes = "A2"
    return ws


# ==============================================================================
# 2. CREATE SCENARIOS SHEET
# ==============================================================================

def create_scenarios_sheet(wb: openpyxl.Workbook):
    """
    Creates the Scenarios tab with Worst / Base / Best case Funds Raised.
    The scenario selection on the Assumptions tab drives which row is active.
    """
    ws = wb.create_sheet("Scenarios")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 36
    for c in ["C", "D", "E", "F", "G"]:
        ws.column_dimensions[c].width = 16
    ws.column_dimensions["H"].width = 40

    # Title
    ws.row_dimensions[1].height = 24
    set_cell(ws, 1, 1, "SCENARIOS — TOTAL FUNDS RAISED", font=TITLE_FONT, fill=SECTION_FILL)
    for c in range(2, 9):
        ws.cell(row=1, column=c).fill = SECTION_FILL

    set_cell(ws, 2, 1,
             "Scenarios only affect Total Funds Raised. All other assumptions remain constant.",
             font=Font(name="Calibri", color="FF0000", size=10, italic=True))
    ws.merge_cells("A2:H2")

    # Column headers
    write_section_header(ws, 4, "SCENARIO DATA TABLE", num_cols=8)
    set_cell(ws, 5, 1, "Scenario",     font=BOLD_BLACK, alignment=CENTER)
    set_cell(ws, 5, 2, "Description",  font=BOLD_BLACK, alignment=LEFT)
    write_year_headers(ws, 5, [2026, 2027, 2028, 2029, 2030])
    set_cell(ws, 5, 8, "Multiplier",   font=BOLD_BLACK, alignment=CENTER)

    # Scenario rows: 6=Worst, 7=Base, 8=Best
    scenario_defs = [
        (6, "Worst", "Base × 0.6  (40% reduction in Funds Raised)", "C19", "0.6"),
        (7, "Base",  "Central assumption — as per business plan",    "C16", "1.0"),
        (8, "Best",  "Base × 1.4  (40% uplift in Funds Raised)",     "C20", "1.4"),
    ]
    base_row = 16   # Assumptions row for base funds raised

    for row_num, scen_name, desc, _mult_ref, mult_label in scenario_defs:
        set_cell(ws, row_num, 1, scen_name, font=BLUE, alignment=CENTER)
        set_cell(ws, row_num, 2, desc,      font=BLACK, alignment=LEFT)
        set_cell(ws, row_num, 8, mult_label, font=BLACK, alignment=CENTER)
        for col_idx in range(5):
            yr_col = col(3 + col_idx)
            if scen_name == "Worst":
                formula = f"=Assumptions!{yr_col}{base_row}*Assumptions!C19"
            elif scen_name == "Base":
                formula = f"=Assumptions!{yr_col}{base_row}"
            else:
                formula = f"=Assumptions!{yr_col}{base_row}*Assumptions!C20"
            set_cell(ws, row_num, 3 + col_idx, formula,
                     font=GREEN, alignment=RIGHT, number_format=GBP_FORMAT)

    # Active scenario row (driven by Assumptions dropdown)
    write_section_header(ws, 10, "ACTIVE SCENARIO (driven by Assumptions!C4 dropdown)")
    set_cell(ws, 11, 1, "Scenario →", font=BOLD_BLACK, alignment=LEFT)
    set_cell(ws, 11, 2, '=Assumptions!C4', font=GREEN, alignment=LEFT)
    set_cell(ws, 12, 1, "Funds Raised", font=BOLD_BLACK, alignment=LEFT)
    set_cell(ws, 12, 2, "(Active)",      font=BLACK,     alignment=LEFT)
    ws.row_dimensions[12].height = 16

    for col_idx in range(5):
        formula = (
            f"=INDEX($C$6:$G$8,"
            f"MATCH(Assumptions!$C$4,$A$6:$A$8,0),"
            f"{col_idx + 1})"
        )
        set_cell(ws, 12, 3 + col_idx, formula,
                 font=GREEN, alignment=RIGHT, number_format=GBP_FORMAT)

    ws.freeze_panes = "A2"
    return ws


# ==============================================================================
# 3. BUILD INCOME STATEMENT
# ==============================================================================

def build_income_statement(wb: openpyxl.Workbook):
    """
    Builds the Income Statement with fully-linked Excel formulas.

    Row map (columns: A=label, B=notes, C=2026, D=2027, E=2028, F=2029, G=2030):
      6  — Total Funds Raised (active scenario)
      7  — Upfront Fee Revenue
      8  — Completion Fee Revenue
      9  — kWh Revenue
      10 — TOTAL GROSS REVENUE
      13 — Smart Contract Deployment & Audits  (COGS)
      14 — Hosting & API Costs (COGS)
      15 — TOTAL COGS
      17 — GROSS PROFIT
      20 — Salaries  (OpEx)
      21 — Marketing & Growth
      22 — Legal & Regulatory
      23 — Insurance
      24 — R&D
      25 — TOTAL OPERATING EXPENSES
      27 — EBIT (Operating Profit / Loss)
      30 — Prior-Year Tax Loss Balance
      31 — Loss Carry-Forward Applied (≤ £5M p.a.)
      32 — Taxable Profit After Loss Relief
      33 — Tax Provision
      34 — Closing Tax Loss Balance (carried forward)
      37 — NET INCOME (After Tax)
    """
    ws = wb.create_sheet("Income Statement")

    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 10
    for c in ["C", "D", "E", "F", "G"]:
        ws.column_dimensions[c].width = 17

    # ── Title ──────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 24
    set_cell(ws, 1, 1, "INCOME STATEMENT", font=TITLE_FONT, fill=SECTION_FILL)
    for c in range(2, 8):
        ws.cell(row=1, column=c).fill = SECTION_FILL

    set_cell(ws, 2, 1, "All figures in GBP (£)", font=BLACK, alignment=LEFT)
    set_cell(ws, 3, 1, "Year →", font=BOLD_BLACK, alignment=LEFT)
    write_year_headers(ws, 3, [2026, 2027, 2028, 2029, 2030])

    # ── Revenue ────────────────────────────────────────────────────────────────
    write_section_header(ws, 5, "REVENUE")

    # Row 6: Funds Raised (active scenario) — cross-sheet link → GREEN
    set_cell(ws, 6, 1, "Total Funds Raised (Active Scenario)", font=GREEN, alignment=LEFT)
    add_comment(ws, 6, 1,
        "Driven by scenario selector on Assumptions tab.\n"
        "References Scenarios!C12:G12 (active scenario row).")
    for col_idx in range(5):
        formula = f"=Scenarios!{col(3 + col_idx)}12"
        set_cell(ws, 6, 3 + col_idx, formula, font=GREEN, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 7: Upfront Fee (2% of Funds Raised)
    set_cell(ws, 7, 1, "Upfront Fee Revenue (2% of Funds Raised)", font=BLACK, alignment=LEFT)
    set_cell(ws, 7, 2, "→ IS", font=BLACK, alignment=CENTER)
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        formula = f"=Assumptions!$C$7*'Income Statement'!{yr_col}6"
        set_cell(ws, 7, 3 + col_idx, formula, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 8: Completion Fee (0.5% spread over 3-year build, straight-line)
    set_cell(ws, 8, 1,
             "Completion Fee Revenue (0.5% over 3-yr build, straight-line)",
             font=BLACK, alignment=LEFT)
    add_comment(ws, 8, 1,
        "Completion fee = 0.5% of project funds raised.\n"
        "Recognised straight-line over 3-year build.\n"
        "Annual recognition: (0.5% / 3) × sum of active cohorts.\n"
        "Year 1: 1/3 × 0.5% × Funds_2026\n"
        "Year 2: 1/3 × 0.5% × (Funds_2026 + Funds_2027)\n"
        "Year 3: 1/3 × 0.5% × (Funds_2026 + Funds_2027 + Funds_2028)\n"
        "Year 4: 1/3 × 0.5% × (Funds_2027 + Funds_2028 + Funds_2029)\n"
        "Year 5: 1/3 × 0.5% × (Funds_2028 + Funds_2029 + Funds_2030)")
    # Columns: C=2026, D=2027, E=2028, F=2029, G=2030
    # Cohort windows (3-year sliding): max(col-2, C) to col
    completion_fee_funds = [
        "C6",
        "C6+D6",
        "C6+D6+E6",
        "D6+E6+F6",
        "E6+F6+G6",
    ]
    for col_idx, funds_expr in enumerate(completion_fee_funds):
        formula = (
            f"=(Assumptions!$C$8/Assumptions!$C$9)"
            f"*('Income Statement'!{funds_expr})"
        )
        set_cell(ws, 8, 3 + col_idx, formula, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 9: kWh Revenue (starts year after completion of Y1 cohort = 2029)
    set_cell(ws, 9, 1,
             "kWh Revenue (2.12% of Funds, 3-yr lag, cumulative cohorts)",
             font=BLACK, alignment=LEFT)
    add_comment(ws, 9, 1,
        "kWh revenue starts the year after project completion (3-year build).\n"
        "Y1 projects complete at end of 2028, so kWh starts in 2029.\n"
        "2026-2028: £0 (no completed projects yet)\n"
        "2029: 2.12% × Funds_2026 (2026 cohort now generating)\n"
        "2030: 2.12% × (Funds_2026 + Funds_2027) (both cohorts generating)\n"
        "kWh contract duration: 10 years. Model shows 5 years only.")
    kwh_formulas = [
        "0",
        "0",
        "0",
        "=Assumptions!$C$11*'Income Statement'!C6",
        "=Assumptions!$C$11*('Income Statement'!C6+'Income Statement'!D6)",
    ]
    for col_idx, formula in enumerate(kwh_formulas):
        set_cell(ws, 9, 3 + col_idx, formula, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 10: Total Gross Revenue
    write_subheader(ws, 10, "TOTAL GROSS REVENUE", num_cols=7)
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        formula = f"=SUM('Income Statement'!{yr_col}7:{yr_col}9)"
        set_cell(ws, 10, 3 + col_idx, formula,
                 font=BOLD_BLACK, fill=SUBSECT_FILL, alignment=RIGHT, number_format=GBP_FORMAT)

    # ── COGS ───────────────────────────────────────────────────────────────────
    write_section_header(ws, 12, "COST OF GOODS SOLD (COGS)")

    # Row 13: Smart Contract costs — cross-sheet link → GREEN
    set_cell(ws, 13, 1, "Smart Contract Deployment & Audits", font=GREEN, alignment=LEFT)
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        formula = f"=Assumptions!{yr_col}24"
        set_cell(ws, 13, 3 + col_idx, formula, font=GREEN, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 14: Hosting & API (1.5% of Revenue)
    set_cell(ws, 14, 1, "Hosting & API Costs (1.5% of Gross Revenue)", font=BLACK, alignment=LEFT)
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        formula = f"=Assumptions!$C$25*'Income Statement'!{yr_col}10"
        set_cell(ws, 14, 3 + col_idx, formula, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 15: Total COGS
    write_subheader(ws, 15, "TOTAL COGS", num_cols=7)
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        formula = f"=SUM('Income Statement'!{yr_col}13:{yr_col}14)"
        set_cell(ws, 15, 3 + col_idx, formula,
                 font=BOLD_BLACK, fill=SUBSECT_FILL, alignment=RIGHT, number_format=GBP_FORMAT)

    # ── Gross Profit ───────────────────────────────────────────────────────────
    write_subheader(ws, 17, "GROSS PROFIT", num_cols=7)
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        formula = f"='Income Statement'!{yr_col}10-'Income Statement'!{yr_col}15"
        set_cell(ws, 17, 3 + col_idx, formula,
                 font=BOLD_BLACK, fill=SUBSECT_FILL, alignment=RIGHT, number_format=GBP_FORMAT)

    # ── Operating Expenses ─────────────────────────────────────────────────────
    write_section_header(ws, 19, "OPERATING EXPENSES")

    opex_defs = [
        # (label, Y1 assump row, Y2-5 pct assumption rows D-G)
        ("Salaries",           28, 36),
        ("Marketing & Growth", 29, 37),
        ("Legal & Regulatory", 30, 38),
        ("Insurance",          31, 39),
        ("R&D",                32, 40),
    ]
    for i, (label, y1_row, pct_row) in enumerate(opex_defs, start=20):
        set_cell(ws, i, 1, label, font=BLACK, alignment=LEFT)
        for col_idx in range(5):
            yr_col = col(3 + col_idx)
            if col_idx == 0:
                # Year 1 (2026): hardcoded ref
                formula = f"=Assumptions!$C${y1_row}"
                set_cell(ws, i, 3, formula, font=GREEN, alignment=RIGHT, number_format=GBP_FORMAT)
            else:
                # Years 2-5: % of gross revenue; assumption col shifts with year
                assump_col = col(4 + col_idx - 1)   # D=2027, E=2028, F=2029, G=2030
                formula = (
                    f"=Assumptions!{assump_col}{pct_row}"
                    f"*'Income Statement'!{yr_col}10"
                )
                set_cell(ws, i, 3 + col_idx, formula,
                         font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 25: Total OpEx
    write_subheader(ws, 25, "TOTAL OPERATING EXPENSES", num_cols=7)
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        formula = f"=SUM('Income Statement'!{yr_col}20:{yr_col}24)"
        set_cell(ws, 25, 3 + col_idx, formula,
                 font=BOLD_BLACK, fill=SUBSECT_FILL, alignment=RIGHT, number_format=GBP_FORMAT)

    # ── EBIT ───────────────────────────────────────────────────────────────────
    write_subheader(ws, 27, "EBIT  (Operating Profit / Loss)", num_cols=7)
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        formula = (
            f"='Income Statement'!{yr_col}17"
            f"-'Income Statement'!{yr_col}25"
        )
        set_cell(ws, 27, 3 + col_idx, formula,
                 font=BOLD_BLACK, fill=SUBSECT_FILL, alignment=RIGHT, number_format=GBP_FORMAT)

    # ── Tax Computation ────────────────────────────────────────────────────────
    write_section_header(ws, 29, "UK CORPORATION TAX COMPUTATION")
    add_comment(ws, 29, 1,
        "UK Corporation Tax: Small Profits Rate 19% (profits ≤ £50K),\n"
        "Main Rate 25% (profits > £250K), marginal relief in between.\n"
        "Tax paid 9 months after year-end (modelled as next-year cash outflow).\n"
        "Loss carry-forward: unlimited duration; £5M annual deduction limit.\n"
        "Source: HMRC CTM04800 / Finance Act 2017.")

    # Row 30: Prior-year loss balance
    set_cell(ws, 30, 1, "Prior-Year Tax Loss Balance (b/f)", font=BLACK, alignment=LEFT)
    # Year 1: no prior year loss
    set_cell(ws, 30, 3, "=0", font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)
    for col_idx in range(1, 5):     # 2027–2030 reference prior-year closing balance (row 34)
        prev_col = col(3 + col_idx - 1)
        formula = f"='Income Statement'!{prev_col}34"
        set_cell(ws, 30, 3 + col_idx, formula, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 31: Loss applied this year (limited to £5M and actual profit)
    set_cell(ws, 31, 1,
             "Loss Carry-Forward Applied (limited to £5M p.a. & actual profit)",
             font=BLACK, alignment=LEFT)
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        # Apply losses only if EBIT > 0; limited to min(prior_loss, 5M limit, EBIT)
        formula = (
            f"=IF('Income Statement'!{yr_col}27>0,"
            f"MIN('Income Statement'!{yr_col}30,"
            f"Assumptions!$C$48,"
            f"'Income Statement'!{yr_col}27),0)"
        )
        set_cell(ws, 31, 3 + col_idx, formula, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 32: Taxable profit after loss relief
    set_cell(ws, 32, 1, "Taxable Profit After Loss Relief", font=BLACK, alignment=LEFT)
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        formula = (
            f"=MAX('Income Statement'!{yr_col}27"
            f"-'Income Statement'!{yr_col}31,0)"
        )
        set_cell(ws, 32, 3 + col_idx, formula, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 33: Tax provision (with marginal relief logic)
    set_cell(ws, 33, 1,
             "Tax Provision (19% / marginal relief / 25%)",
             font=BLACK, alignment=LEFT)
    add_comment(ws, 33, 1,
        "UK Corp Tax with marginal relief:\n"
        "  ≤ £50K: 19% × Profit\n"
        "  £50K–£250K: 25% × P − (£250K − P) × 3/200\n"
        "  > £250K: 25% × Profit")
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        # Nested IF for marginal relief; C$40=LPT(50K), C$41=UPT(250K)
        formula = (
            f"=IF('Income Statement'!{yr_col}32<=0,0,"
            f"IF('Income Statement'!{yr_col}32<=Assumptions!$C$45,"
            f"Assumptions!$C$43*'Income Statement'!{yr_col}32,"
            f"IF('Income Statement'!{yr_col}32<=Assumptions!$C$46,"
            f"Assumptions!$C$44*'Income Statement'!{yr_col}32"
            f"-(Assumptions!$C$46-'Income Statement'!{yr_col}32)*3/200,"
            f"Assumptions!$C$44*'Income Statement'!{yr_col}32)))"
        )
        set_cell(ws, 33, 3 + col_idx, formula, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 34: Closing tax loss balance (carried forward)
    set_cell(ws, 34, 1,
             "Closing Tax Loss Balance (carried forward to next year)",
             font=BLACK, alignment=LEFT)
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        # Remaining prior loss + new losses generated this year (if EBIT < 0)
        formula = (
            f"=MAX(0,'Income Statement'!{yr_col}30"
            f"-'Income Statement'!{yr_col}31)"
            f"+MAX(0,-'Income Statement'!{yr_col}27)"
        )
        set_cell(ws, 34, 3 + col_idx, formula, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    # ── Net Income ─────────────────────────────────────────────────────────────
    write_subheader(ws, 36, "NET INCOME  (After Tax)", num_cols=7)
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        formula = (
            f"='Income Statement'!{yr_col}27"
            f"-'Income Statement'!{yr_col}33"
        )
        c = set_cell(ws, 36, 3 + col_idx, formula,
                     font=BOLD_BLACK, fill=SUBSECT_FILL, alignment=RIGHT, number_format=GBP_FORMAT)
        c.border = Border(top=THIN, bottom=THIN)

    ws.freeze_panes = "C4"
    return ws


# ==============================================================================
# 4. BUILD BALANCE SHEET
# ==============================================================================

def build_balance_sheet(wb: openpyxl.Workbook):
    """
    Builds the Balance Sheet fully linked to the Income Statement and Cash Flow.

    Row map:
      6  — Cash & Cash Equivalents       (= Cash Flow closing cash)
      7  — Accounts Receivable           (= Revenue × DSO / 365)
      8  — TOTAL CURRENT ASSETS
      9  — Total Assets
      12 — Accounts Payable              (= (COGS+OpEx) × DPO / 365)
      13 — Corporation Tax Payable       (= current year tax provision, 9-month lag)
      14 — TOTAL CURRENT LIABILITIES
      15 — Total Liabilities
      18 — Share Capital (cumulative equity raises)
      19 — Retained Earnings (cumulative net income)
      20 — TOTAL EQUITY
      22 — BALANCE CHECK  (Assets − Liabilities − Equity = 0)
    """
    ws = wb.create_sheet("Balance Sheet")

    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 10
    for c in ["C", "D", "E", "F", "G"]:
        ws.column_dimensions[c].width = 17

    # ── Title ──────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 24
    set_cell(ws, 1, 1, "BALANCE SHEET", font=TITLE_FONT, fill=SECTION_FILL)
    for c in range(2, 8):
        ws.cell(row=1, column=c).fill = SECTION_FILL

    set_cell(ws, 2, 1, "All figures in GBP (£)", font=BLACK)
    set_cell(ws, 3, 1, "Year →", font=BOLD_BLACK)
    write_year_headers(ws, 3, [2026, 2027, 2028, 2029, 2030])

    # ── ASSETS ────────────────────────────────────────────────────────────────
    write_section_header(ws, 5, "ASSETS")

    # Row 6: Cash (linked to Cash Flow closing cash, row 23)
    set_cell(ws, 6, 1, "Cash & Cash Equivalents", font=GREEN, alignment=LEFT)
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        formula = f"='Cash Flow'!{yr_col}23"
        set_cell(ws, 6, 3 + col_idx, formula, font=GREEN, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 7: Accounts Receivable
    set_cell(ws, 7, 1,
             "Accounts Receivable  (Revenue × DSO / 365)",
             font=BLACK, alignment=LEFT)
    add_comment(ws, 7, 1, f"DSO = 45 days (from Assumptions!C55)")
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        formula = (
            f"='Income Statement'!{yr_col}10"
            f"*Assumptions!$C$55/365"
        )
        set_cell(ws, 7, 3 + col_idx, formula, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 8: Total Current Assets
    write_subheader(ws, 8, "TOTAL CURRENT ASSETS", num_cols=7)
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        formula = f"=SUM('Balance Sheet'!{yr_col}6:{yr_col}7)"
        set_cell(ws, 8, 3 + col_idx, formula,
                 font=BOLD_BLACK, fill=SUBSECT_FILL, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 9: Total Assets (no fixed assets — asset-light)
    write_subheader(ws, 9, "TOTAL ASSETS", num_cols=7)
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        formula = f"='Balance Sheet'!{yr_col}8"
        set_cell(ws, 9, 3 + col_idx, formula,
                 font=BOLD_BLACK, fill=SUBSECT_FILL, alignment=RIGHT, number_format=GBP_FORMAT,
                 border=Border(bottom=THIN))

    # ── LIABILITIES ───────────────────────────────────────────────────────────
    write_section_header(ws, 11, "LIABILITIES")

    # Row 12: Accounts Payable
    set_cell(ws, 12, 1,
             "Accounts Payable  ((COGS + OpEx) × DPO / 365)",
             font=BLACK, alignment=LEFT)
    add_comment(ws, 12, 1, "DPO = 30 days (from Assumptions!C56)")
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        formula = (
            f"=('Income Statement'!{yr_col}15"
            f"+'Income Statement'!{yr_col}25)"
            f"*Assumptions!$C$56/365"
        )
        set_cell(ws, 12, 3 + col_idx, formula, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 13: Corporation Tax Payable (9-month lag — paid next year)
    set_cell(ws, 13, 1,
             "Corporation Tax Payable  (current year provision, paid next year)",
             font=BLACK, alignment=LEFT)
    add_comment(ws, 13, 1,
        "Tax paid 9 months after year-end.\n"
        "Current year tax provision = liability on Balance Sheet.\n"
        "Cash payment shown as prior-year outflow in Cash Flow.")
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        formula = f"='Income Statement'!{yr_col}33"
        set_cell(ws, 13, 3 + col_idx, formula, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 14: Total Current Liabilities
    write_subheader(ws, 14, "TOTAL CURRENT LIABILITIES", num_cols=7)
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        formula = f"=SUM('Balance Sheet'!{yr_col}12:{yr_col}13)"
        set_cell(ws, 14, 3 + col_idx, formula,
                 font=BOLD_BLACK, fill=SUBSECT_FILL, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 15: Total Liabilities
    write_subheader(ws, 15, "TOTAL LIABILITIES", num_cols=7)
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        formula = f"='Balance Sheet'!{yr_col}14"
        set_cell(ws, 15, 3 + col_idx, formula,
                 font=BOLD_BLACK, fill=SUBSECT_FILL, alignment=RIGHT, number_format=GBP_FORMAT,
                 border=Border(bottom=THIN))

    # ── EQUITY ────────────────────────────────────────────────────────────────
    write_section_header(ws, 17, "EQUITY")

    # Row 18: Share Capital (cumulative equity raises)
    set_cell(ws, 18, 1,
             "Share Capital  (cumulative equity raises to date)",
             font=GREEN, alignment=LEFT)
    # Year 1: just the Y1 equity raise
    formula_y1 = "=Assumptions!C52"
    set_cell(ws, 18, 3, formula_y1, font=GREEN, alignment=RIGHT, number_format=GBP_FORMAT)
    # Years 2-5: prior year share capital + current year raise
    for col_idx in range(1, 5):
        yr_col  = col(3 + col_idx)
        prev_col = col(3 + col_idx - 1)
        raise_col = col(3 + col_idx)   # same column in Assumptions row 52
        formula = (
            f"='Balance Sheet'!{prev_col}18"
            f"+Assumptions!{raise_col}52"
        )
        set_cell(ws, 18, 3 + col_idx, formula, font=GREEN, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 19: Retained Earnings (cumulative net income)
    set_cell(ws, 19, 1,
             "Retained Earnings  (cumulative net income)",
             font=BLACK, alignment=LEFT)
    # Year 1: just Y1 net income
    set_cell(ws, 19, 3,
             "='Income Statement'!C36",
             font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)
    for col_idx in range(1, 5):
        yr_col  = col(3 + col_idx)
        prev_col = col(3 + col_idx - 1)
        formula = (
            f"='Balance Sheet'!{prev_col}19"
            f"+'Income Statement'!{yr_col}36"
        )
        set_cell(ws, 19, 3 + col_idx, formula, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 20: Total Equity
    write_subheader(ws, 20, "TOTAL EQUITY", num_cols=7)
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        formula = f"=SUM('Balance Sheet'!{yr_col}18:{yr_col}19)"
        set_cell(ws, 20, 3 + col_idx, formula,
                 font=BOLD_BLACK, fill=SUBSECT_FILL, alignment=RIGHT, number_format=GBP_FORMAT,
                 border=Border(bottom=THIN))

    # ── Balance Check ─────────────────────────────────────────────────────────
    write_section_header(ws, 22, "BALANCE CHECK  (Total Assets − Total Liabilities − Total Equity = 0)")
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        formula = (
            f"='Balance Sheet'!{yr_col}9"
            f"-'Balance Sheet'!{yr_col}15"
            f"-'Balance Sheet'!{yr_col}20"
        )
        cell = set_cell(ws, 22, 3 + col_idx, formula,
                        font=BOLD_BLACK, alignment=CENTER, number_format=GBP_FORMAT)
        # Conditional-style comment only; Excel CF would require VBA
        cell.border = BOX_BORDER

    add_comment(ws, 22, 3,
        "Balance Check = Total Assets − Total Liabilities − Total Equity.\n"
        "Should equal £0 in every column.\n"
        "Non-zero indicates a formula error — check Cash Flow linkage.")

    ws.freeze_panes = "C4"
    return ws


# ==============================================================================
# 5. BUILD CASH FLOW STATEMENT
# ==============================================================================

def build_cash_flow(wb: openpyxl.Workbook):
    """
    Builds the Cash Flow Statement (indirect method) fully linked.

    Row map:
      6  — Net Income (from IS row 36)
      7  — Change in Accounts Receivable
      8  — Change in Accounts Payable
      9  — Change in Corporation Tax Payable (current provision − prior payable)
      10 — NET CASH FROM OPERATIONS
      13 — Capital Expenditure
      14 — NET CASH FROM INVESTING
      17 — Equity Raised
      18 — NET CASH FROM FINANCING
      21 — Net Change in Cash
      22 — Opening Cash
      23 — CLOSING CASH  (→ Balance Sheet Cash row 6)

    Note on tax timing:
      Net Income already includes current-year tax expense.
      "Change in Tax Payable" adds back the current-year provision (not yet cash)
      and deducts the prior-year provision that has now been paid.
      This is equivalent to the "Tax Paid (prior year)" presentation but
      ensures the Balance Sheet always balances.
    """
    ws = wb.create_sheet("Cash Flow")

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 10
    for c in ["C", "D", "E", "F", "G"]:
        ws.column_dimensions[c].width = 17

    # ── Title ──────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 24
    set_cell(ws, 1, 1, "CASH FLOW STATEMENT  (Indirect Method)", font=TITLE_FONT, fill=SECTION_FILL)
    for c in range(2, 8):
        ws.cell(row=1, column=c).fill = SECTION_FILL

    set_cell(ws, 2, 1, "All figures in GBP (£)", font=BLACK)
    set_cell(ws, 3, 1, "Year →", font=BOLD_BLACK)
    write_year_headers(ws, 3, [2026, 2027, 2028, 2029, 2030])

    # ── Operating Activities ───────────────────────────────────────────────────
    write_section_header(ws, 5, "OPERATING ACTIVITIES")

    # Row 6: Net Income
    set_cell(ws, 6, 1, "Net Income (after tax)", font=GREEN, alignment=LEFT)
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        formula = f"='Income Statement'!{yr_col}36"
        set_cell(ws, 6, 3 + col_idx, formula, font=GREEN, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 7: Change in Accounts Receivable (increase = outflow → negative)
    set_cell(ws, 7, 1,
             "Change in Accounts Receivable  (increase = cash outflow)",
             font=BLACK, alignment=LEFT)
    add_comment(ws, 7, 1,
        "Cash impact: −(Current AR − Prior AR)\n"
        "Increase in AR = customers owe more = cash not yet received.")
    # Year 1: prior AR = 0
    formula_y1 = "=-('Balance Sheet'!C7-0)"
    set_cell(ws, 7, 3, formula_y1, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)
    for col_idx in range(1, 5):
        yr_col  = col(3 + col_idx)
        prev_col = col(3 + col_idx - 1)
        formula = (
            f"=-('Balance Sheet'!{yr_col}7"
            f"-'Balance Sheet'!{prev_col}7)"
        )
        set_cell(ws, 7, 3 + col_idx, formula, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 8: Change in Accounts Payable (increase = inflow → positive)
    set_cell(ws, 8, 1,
             "Change in Accounts Payable  (increase = cash inflow)",
             font=BLACK, alignment=LEFT)
    add_comment(ws, 8, 1,
        "Cash impact: Current AP − Prior AP\n"
        "Increase in AP = owing more to suppliers = cash retained.")
    # Year 1: prior AP = 0
    formula_y1 = "='Balance Sheet'!C12-0"
    set_cell(ws, 8, 3, formula_y1, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)
    for col_idx in range(1, 5):
        yr_col  = col(3 + col_idx)
        prev_col = col(3 + col_idx - 1)
        formula = (
            f"='Balance Sheet'!{yr_col}12"
            f"-'Balance Sheet'!{prev_col}12"
        )
        set_cell(ws, 8, 3 + col_idx, formula, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 9: Change in Corporation Tax Payable
    # = Current year tax provision (not yet paid) − Prior year tax provision (now paid in cash)
    set_cell(ws, 9, 1,
             "Change in Corporation Tax Payable  (current provision − prior paid)",
             font=BLACK, alignment=LEFT)
    add_comment(ws, 9, 1,
        "Indirect method tax adjustment:\n"
        "+ Current year tax provision  (non-cash deduction — not yet paid)\n"
        "− Prior year tax provision  (now settled in cash, 9-month lag)\n"
        "Net = change in Tax Payable balance.\n"
        "This ensures the Balance Sheet balances correctly.")
    # Year 1: prior tax = 0, so just +current tax
    formula_y1 = "='Balance Sheet'!C13-0"
    set_cell(ws, 9, 3, formula_y1, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)
    for col_idx in range(1, 5):
        yr_col  = col(3 + col_idx)
        prev_col = col(3 + col_idx - 1)
        formula = (
            f"='Balance Sheet'!{yr_col}13"
            f"-'Balance Sheet'!{prev_col}13"
        )
        set_cell(ws, 9, 3 + col_idx, formula, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 10: Net Cash from Operations
    write_subheader(ws, 10, "NET CASH FROM OPERATIONS", num_cols=7)
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        formula = f"=SUM('Cash Flow'!{yr_col}6:{yr_col}9)"
        set_cell(ws, 10, 3 + col_idx, formula,
                 font=BOLD_BLACK, fill=SUBSECT_FILL, alignment=RIGHT, number_format=GBP_FORMAT)

    # ── Investing Activities ───────────────────────────────────────────────────
    write_section_header(ws, 12, "INVESTING ACTIVITIES")

    # Row 13: CapEx (£0 — asset-light)
    set_cell(ws, 13, 1,
             "Capital Expenditure / CapEx  (asset-light — £0)",
             font=GREEN, alignment=LEFT)
    for col_idx in range(5):
        formula = "=-Assumptions!$C$62"
        set_cell(ws, 13, 3 + col_idx, formula, font=GREEN, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 14: Net Cash from Investing
    write_subheader(ws, 14, "NET CASH FROM INVESTING", num_cols=7)
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        formula = f"='Cash Flow'!{yr_col}13"
        set_cell(ws, 14, 3 + col_idx, formula,
                 font=BOLD_BLACK, fill=SUBSECT_FILL, alignment=RIGHT, number_format=GBP_FORMAT)

    # ── Financing Activities ───────────────────────────────────────────────────
    write_section_header(ws, 16, "FINANCING ACTIVITIES")

    # Row 17: Equity Raised
    set_cell(ws, 17, 1,
             "Equity Raised  (per Assumptions)",
             font=GREEN, alignment=LEFT)
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        formula = f"=Assumptions!{yr_col}52"
        set_cell(ws, 17, 3 + col_idx, formula, font=GREEN, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 18: Net Cash from Financing
    write_subheader(ws, 18, "NET CASH FROM FINANCING", num_cols=7)
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        formula = f"='Cash Flow'!{yr_col}17"
        set_cell(ws, 18, 3 + col_idx, formula,
                 font=BOLD_BLACK, fill=SUBSECT_FILL, alignment=RIGHT, number_format=GBP_FORMAT)

    # ── Cash Summary ───────────────────────────────────────────────────────────
    write_section_header(ws, 20, "CASH SUMMARY")

    # Row 21: Net Change in Cash
    set_cell(ws, 21, 1, "Net Change in Cash", font=BLACK, alignment=LEFT)
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        formula = (
            f"='Cash Flow'!{yr_col}10"
            f"+'Cash Flow'!{yr_col}14"
            f"+'Cash Flow'!{yr_col}18"
        )
        set_cell(ws, 21, 3 + col_idx, formula, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 22: Opening Cash
    set_cell(ws, 22, 1, "Opening Cash", font=BLACK, alignment=LEFT)
    # Year 1: from Assumptions (£0 before equity raise; equity is in Financing above)
    formula_y1 = "=Assumptions!$C$59"
    set_cell(ws, 22, 3, formula_y1, font=GREEN, alignment=RIGHT, number_format=GBP_FORMAT)
    for col_idx in range(1, 5):
        yr_col  = col(3 + col_idx)
        prev_col = col(3 + col_idx - 1)
        formula = f"='Cash Flow'!{prev_col}23"
        set_cell(ws, 22, 3 + col_idx, formula, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 23: Closing Cash (links to Balance Sheet)
    write_subheader(ws, 23, "CLOSING CASH  (= Balance Sheet: Cash & Cash Equivalents)", num_cols=7)
    for col_idx in range(5):
        yr_col = col(3 + col_idx)
        formula = f"='Cash Flow'!{yr_col}21+'Cash Flow'!{yr_col}22"
        set_cell(ws, 23, 3 + col_idx, formula,
                 font=BOLD_BLACK, fill=SUBSECT_FILL, alignment=RIGHT, number_format=GBP_FORMAT,
                 border=Border(top=THIN, bottom=THIN))

    ws.freeze_panes = "C4"
    return ws


# ==============================================================================
# 6. CREATE VALUATION SHEET
# ==============================================================================

def create_valuation_sheet(wb: openpyxl.Workbook, assumptions: dict):
    """
    Builds the Valuation tab — a VC valuation model that:
      - Projects revenue from Income Statement (years 1-5) then grows at 60% CAGR
      - Applies an EV/Revenue multiple to derive Enterprise Value per year
      - Discounts equity values back using tiered discount rates (50%/35%/25%)
      - Calculates investor IRR and cash-on-cash multiple

    Row map:
      3   — Section A header: VALUATION ASSUMPTIONS
      4   — Currency
      5   — Valuation date
      6   — Exit year
      7   — EV/Revenue multiple
      8   — Financial debt at exit
      9   — Cash at exit (linked to Balance Sheet G6)
      10  — Required return Y1-3
      11  — Required return Y4-6
      12  — Required return Y7-10
      13  — Revenue CAGR Y5-Y10
      14  — Equity growth rate Y5-Y10
      16  — Year headers (0-10)
      17  — Section B header: REVENUE PROJECTIONS
      18  — Net Revenue (years 1-5 from IS, years 6-10 at 60% CAGR)
      20  — Section C header: VC VALUATION TABLE
      21  — Forecast year (0-10)
      22  — Cash flow dates (EDATE 12-month increments)
      23  — Net Revenue
      24  — Enterprise Value (EV = Revenue × multiple)
      25  — Financial Debt (only at exit year)
      26  — Cash (exit year from BS; years 6-10 at 2% revenue; else 0)
      27  — Equity Value
      28  — Discount Rate (tiered)
      29  — Discount Period (cumulative years from valuation date)
      30  — Discount Factor
      31  — Present Value (exit year onwards)
      33  — Section D header: INVESTOR CALCULATION
      34  — Investment amount
      35  — Equity stake at entry
      36  — Dilution effect
      37  — Equity stake at exit
      38  — Exit proceeds (equity_value[exit_year] × stake_at_exit)
      39  — Investor cash flows (year 0: -investment, year 5: exit proceeds)
      40  — IRR (=IRR on cash flow range)
      41  — Cash-on-cash multiple
      43  — Section E header: SUMMARY METRICS
      44  — Equity Value Today (PV at exit year)
      45  — Investor IRR
      46  — Cash-on-Cash Multiple
      47  — Exit Proceeds
    """
    ws = wb.create_sheet("Valuation")

    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 15
    for c_idx in range(3, 15):          # columns C through N (year 0 through 10)
        ws.column_dimensions[col(c_idx)].width = 15

    # ── Title ──────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 24
    set_cell(ws, 1, 1, "VALUATION MODEL — VC METHOD", font=TITLE_FONT, fill=SECTION_FILL)
    for c in range(2, 15):
        ws.cell(row=1, column=c).fill = SECTION_FILL

    set_cell(ws, 2, 1, "All figures in GBP (£) unless stated", font=BLACK, alignment=LEFT)

    # ── Section A: Valuation Assumptions ──────────────────────────────────────
    write_section_header(ws, 3, "A. VALUATION ASSUMPTIONS", num_cols=14)

    assump_rows = [
        (4,  "Currency",                           "GBP",                                       None),
        (5,  "Valuation Date",                      assumptions["valuation_date"],               None),
        (6,  "Exit Year",                           assumptions["exit_year"],                    NUM_FORMAT),
        (7,  "EV / Revenue Multiple",               assumptions["ev_revenue_multiple"],          '0"x"'),
        (8,  "Financial Debt at Exit (£)",          assumptions["financial_debt_at_exit"],       GBP_FORMAT),
        (9,  "Cash at Exit (£)  [→ Balance Sheet G6]", "='Balance Sheet'!G6",                   GBP_FORMAT),
        (10, "Required Return Y1-3",                assumptions["required_return_y1_3"],         PCT_FORMAT),
        (11, "Required Return Y4-6",                assumptions["required_return_y3_6"],         PCT_FORMAT),
        (12, "Required Return Y7-10",               assumptions["required_return_y6_10"],        PCT_FORMAT),
        (13, "Revenue CAGR Y5-Y10",                 assumptions["revenue_cagr_y5_y10"],          PCT_FORMAT),
        (14, "Equity Growth Rate Y5-Y10",           assumptions["equity_growth_rate_y5_10"],     PCT_FORMAT),
    ]
    for row_num, label, val, fmt in assump_rows:
        set_cell(ws, row_num, 1, label, font=BLACK, alignment=LEFT)
        if isinstance(val, str) and val.startswith("="):
            fnt = GREEN
        elif isinstance(val, (int, float)):
            fnt = BLUE
        else:
            fnt = BLUE
        c = set_cell(ws, row_num, 3, val, font=fnt, alignment=CENTER,
                     number_format=fmt if fmt else "@")

    add_comment(ws, 10, 3,
        "Tiered discount rates (FAST VC convention):\n"
        "  Y1-3:  50% (highest risk — early stage)\n"
        "  Y4-6:  35% (growth stage)\n"
        "  Y7-10: 25% (more mature, lower risk)\n"
        "Discount is compounded through each tier cumulatively.")

    # ── Year header row (0-10 across columns C-M) ─────────────────────────────
    set_cell(ws, 16, 1, "Forecast Year →", font=BOLD_BLACK, alignment=LEFT)
    for yr in range(11):   # 0 to 10
        set_cell(ws, 16, 3 + yr, yr, font=BOLD_BLACK, alignment=CENTER, border=THIN_BORDER)

    # ── Section B: Revenue Projections ────────────────────────────────────────
    write_section_header(ws, 17, "B. REVENUE PROJECTIONS  (£)", num_cols=14)

    set_cell(ws, 18, 1, "Net Revenue (Total Gross Revenue)", font=BLACK, alignment=LEFT)
    set_cell(ws, 18, 3, 0, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)   # year 0: no revenue

    # Years 1-5: link to Income Statement row 10 (Total Gross Revenue), columns C-G
    is_cols = ["C", "D", "E", "F", "G"]
    for i, is_col in enumerate(is_cols):
        formula = f"='Income Statement'!{is_col}10"
        set_cell(ws, 18, 4 + i, formula, font=GREEN, alignment=RIGHT, number_format=GBP_FORMAT)

    # Years 6-10: grow at revenue_cagr (60%) from prior year
    # Year 6 = G18 * (1 + cagr); year 7 = H18 * (1+cagr) etc.
    for i in range(5):   # columns I through M (indices 9-13, years 6-10)
        prev_col_letter = col(3 + 5 + i)   # H, I, J, K, L
        curr_col_idx = 3 + 5 + i + 1       # I, J, K, L, M
        formula = f"={prev_col_letter}18*(1+$C$13)"
        set_cell(ws, 18, curr_col_idx, formula, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    add_comment(ws, 18, 4,
        "Years 1-5: linked directly to 'Income Statement'!C10:G10 (Total Gross Revenue).\n"
        "Years 6-10: prior year revenue × (1 + Revenue CAGR Y5-Y10) = 60% growth.")

    # ── Section C: VC Valuation Table ─────────────────────────────────────────
    write_section_header(ws, 20, "C. VC VALUATION TABLE", num_cols=14)

    # Row 21: Forecast Year (0-10)
    set_cell(ws, 21, 1, "Forecast Year", font=BLACK, alignment=LEFT)
    for yr in range(11):
        set_cell(ws, 21, 3 + yr, yr, font=BLACK, alignment=CENTER, number_format=NUM_FORMAT)

    # Row 22: Cash flow dates — EDATE from valuation_date in 12-month increments
    # valuation_date is in C5 (row 5, col 3)
    set_cell(ws, 22, 1, "Cash Flow Date", font=BLACK, alignment=LEFT)
    # Year 0 = valuation date
    set_cell(ws, 22, 3, "=$C$5", font=BLACK, alignment=CENTER, number_format="YYYY-MM-DD")
    for yr in range(1, 11):
        prev_col_letter = col(3 + yr - 1)
        formula = f"=EDATE({prev_col_letter}22,12)"
        set_cell(ws, 22, 3 + yr, formula, font=BLACK, alignment=CENTER, number_format="YYYY-MM-DD")

    # Row 23: Net Revenue (mirror of row 18)
    set_cell(ws, 23, 1, "Net Revenue (£)", font=BLACK, alignment=LEFT)
    for c_idx in range(11):
        src_col_letter = col(3 + c_idx)
        formula = f"={src_col_letter}18"
        set_cell(ws, 23, 3 + c_idx, formula, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 24: Enterprise Value = Revenue × multiple
    set_cell(ws, 24, 1, "Enterprise Value  (EV = Revenue × Multiple)", font=BLACK, alignment=LEFT)
    for c_idx in range(11):
        src_col_letter = col(3 + c_idx)
        formula = f"={src_col_letter}23*$C$7"
        set_cell(ws, 24, 3 + c_idx, formula, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 25: Financial Debt (negative; only at exit year, i.e. col H = year 5)
    # exit_year = C6 (row 6, col 3)
    set_cell(ws, 25, 1, "Financial Debt at Exit (£)", font=BLACK, alignment=LEFT)
    for c_idx in range(11):
        yr_num = c_idx  # 0-10
        if yr_num == 0:
            set_cell(ws, 25, 3 + c_idx, 0, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)
        else:
            src_col_letter = col(3 + c_idx)
            # IF forecast_year == exit_year: -debt, else 0
            formula = f"=IF({src_col_letter}21=$C$6,-$C$8,0)"
            set_cell(ws, 25, 3 + c_idx, formula, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 26: Cash
    # At exit year: link to Balance Sheet G6 (cash at year 5 = col G)
    # Years > exit: 2% of revenue
    # Else: 0
    set_cell(ws, 26, 1, "Cash at Exit / 2% of Revenue (£)", font=BLACK, alignment=LEFT)
    add_comment(ws, 26, 1,
        "Cash logic by year:\n"
        "  Exit year (5): linked to 'Balance Sheet'!G6 (closing cash 2030)\n"
        "  Years 6-10: 2% of net revenue (proxy for cash balance)\n"
        "  All other years: £0")
    for c_idx in range(11):
        yr_num = c_idx
        if yr_num == 0:
            set_cell(ws, 26, 3 + c_idx, 0, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)
        else:
            src_col_letter = col(3 + c_idx)
            formula = (
                f"=IF({src_col_letter}21=$C$6,$C$9,"
                f"IF({src_col_letter}21>$C$6,{src_col_letter}23*0.02,0))"
            )
            fnt = GREEN if yr_num == 5 else BLACK  # year 5 references BS via C9
            set_cell(ws, 26, 3 + c_idx, formula, font=fnt, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 27: Equity Value
    # Years 1-exit: EV + debt + cash
    # Years > exit: prior year equity value × (1 + equity_growth_rate)
    set_cell(ws, 27, 1, "Equity Value (£)", font=BLACK, alignment=LEFT)
    add_comment(ws, 27, 1,
        "Equity Value calculation:\n"
        "  Years 1-5 (up to exit): EV + Financial Debt + Cash\n"
        "  Years 6-10 (post-exit): prior year Equity Value × (1 + Equity Growth Rate)")
    # Year 0: no equity value
    set_cell(ws, 27, 3, 0, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)
    for c_idx in range(1, 11):
        yr_num = c_idx
        src_col_letter = col(3 + c_idx)
        prev_col_letter = col(3 + c_idx - 1)
        formula = (
            f"=IF({src_col_letter}21<=$C$6,"
            f"{src_col_letter}24+{src_col_letter}25+{src_col_letter}26,"
            f"{prev_col_letter}27*(1+$C$14))"
        )
        set_cell(ws, 27, 3 + c_idx, formula, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 28: Discount Rate (tiered by year)
    set_cell(ws, 28, 1, "Discount Rate (tiered)", font=BLACK, alignment=LEFT)
    add_comment(ws, 28, 1,
        "Tiered discount rate:\n"
        "  Years 1-3:  50% (C10)\n"
        "  Years 4-6:  35% (C11)\n"
        "  Years 7-10: 25% (C12)")
    # Year 0: no discount rate
    set_cell(ws, 28, 3, 0, font=BLACK, alignment=RIGHT, number_format=PCT_FORMAT)
    for c_idx in range(1, 11):
        yr_num = c_idx
        src_col_letter = col(3 + c_idx)
        formula = (
            f"=IF({src_col_letter}21<=3,$C$10,"
            f"IF({src_col_letter}21<=6,$C$11,$C$12))"
        )
        set_cell(ws, 28, 3 + c_idx, formula, font=BLACK, alignment=RIGHT, number_format=PCT_FORMAT)

    # Row 29: Discount Period (cumulative years from valuation date)
    set_cell(ws, 29, 1, "Discount Period (years from valuation date)", font=BLACK, alignment=LEFT)
    # Year 0: period = 0
    set_cell(ws, 29, 3, 0, font=BLACK, alignment=RIGHT, number_format="0.0")
    for c_idx in range(1, 11):
        src_col_letter = col(3 + c_idx)
        # ROUND((date - valuation_date)/365, 1) but cumulative
        # = C29 + ROUND((this_date - prev_date)/365, 1)
        prev_col_letter = col(3 + c_idx - 1)
        formula = f"={prev_col_letter}29+ROUND(({src_col_letter}22-{prev_col_letter}22)/365,1)"
        set_cell(ws, 29, 3 + c_idx, formula, font=BLACK, alignment=RIGHT, number_format="0.0")

    # Row 30: Discount Factor = 1 / (1 + rate)^period
    set_cell(ws, 30, 1, "Discount Factor", font=BLACK, alignment=LEFT)
    set_cell(ws, 30, 3, 1, font=BLACK, alignment=RIGHT, number_format="0.0000")
    for c_idx in range(1, 11):
        src_col_letter = col(3 + c_idx)
        formula = f"=1/(1+{src_col_letter}28)^{src_col_letter}29"
        set_cell(ws, 30, 3 + c_idx, formula, font=BLACK, alignment=RIGHT, number_format="0.0000")

    # Row 31: Present Value — compound discount through tiers, exit year onwards only
    # PV formula for year n: equity_value / ((1+r_y1_3)^min(n,3) * (1+r_y3_6)^max(0,min(n,6)-3) * (1+r_y6_10)^max(0,n-6))
    set_cell(ws, 31, 1, "Present Value (£) — exit year onwards", font=BLACK, alignment=LEFT)
    add_comment(ws, 31, 1,
        "Present Value uses compound tiered discounting:\n"
        "  PV = Equity Value / ((1+r_y1_3)^y1_3_yrs × (1+r_y3_6)^y3_6_yrs × (1+r_y6_10)^y6_10_yrs)\n"
        "  Only calculated from exit year (5) onwards.\n"
        "  y1_3_yrs = MIN(n,3)\n"
        "  y3_6_yrs = MAX(0, MIN(n,6)-3)\n"
        "  y6_10_yrs = MAX(0, n-6)")
    # Years 0-4: blank/0
    for c_idx in range(5):
        set_cell(ws, 31, 3 + c_idx, "", font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)
    # Years 5-10: compound discount
    for c_idx in range(5, 11):
        yr_num = c_idx
        src_col_letter = col(3 + c_idx)
        # For year n: y1_3_yrs = MIN(n,3), y3_6_yrs = MAX(0,MIN(n,6)-3), y6_10_yrs = MAX(0,n-6)
        formula = (
            f"={src_col_letter}27/"
            f"((1+$C$10)^MIN({yr_num},3)"
            f"*(1+$C$11)^MAX(0,MIN({yr_num},6)-3)"
            f"*(1+$C$12)^MAX(0,{yr_num}-6))"
        )
        set_cell(ws, 31, 3 + c_idx, formula, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    # ── Section D: Investor Calculation ───────────────────────────────────────
    write_section_header(ws, 33, "D. INVESTOR CALCULATION", num_cols=14)

    # Labels in column A, values in column C (inputs), then year-by-year in C-M
    set_cell(ws, 34, 1, "Investment Amount (£)",          font=BLACK, alignment=LEFT)
    set_cell(ws, 34, 3, 250_000, font=BLUE, alignment=RIGHT, number_format=GBP_FORMAT)

    set_cell(ws, 35, 1, "Equity Stake at Entry",          font=BLACK, alignment=LEFT)
    # Pull from Cap Table: Pre-seed T1 ownership % (Section B, row 9, col C)
    set_cell(ws, 35, 3, "='Cap Table'!C9", font=GREEN, alignment=CENTER, number_format=PCT_FORMAT)
    add_comment(ws, 35, 3, "Linked to Cap Table!C9 — Pre-seed T1 ownership % granted at round.")

    set_cell(ws, 36, 1, "Dilution Effect",                font=BLACK, alignment=LEFT)
    # Dilution = final ownership - entry ownership (from Cap Table Section D)
    set_cell(ws, 36, 3, "='Cap Table'!C30-'Cap Table'!C9", font=GREEN, alignment=CENTER, number_format=PCT_FORMAT)
    add_comment(ws, 36, 3,
        "Dilution = Pre-seed T1 final ownership % (Section D, row 30) minus initial stake (Section B, row 9).\n"
        "Negative value = dilution from subsequent funding rounds.")

    set_cell(ws, 37, 1, "Equity Stake at Exit",           font=BLACK, alignment=LEFT)
    set_cell(ws, 37, 3, "=$C$35+$C$36", font=BLACK, alignment=CENTER, number_format=PCT_FORMAT)

    set_cell(ws, 38, 1, "Exit Proceeds  (Equity Value × Stake at Exit)", font=BLACK, alignment=LEFT)
    # Exit proceeds = equity value at exit year × stake at exit
    # Exit year is year 5 → column H (3+5=8)
    exit_col_letter = col(3 + 5)   # H
    set_cell(ws, 38, 3, f"={exit_col_letter}27*$C$37", font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 39: Investor Cash Flows across years 0-10
    set_cell(ws, 39, 1, "Investor Cash Flows (£)", font=BLACK, alignment=LEFT)
    for c_idx in range(11):
        yr_num = c_idx
        src_col_letter = col(3 + c_idx)
        if yr_num == 0:
            formula = "=-$C$34"     # outflow at year 0
        elif yr_num == 5:           # exit year
            formula = f"=$C$38"
        else:
            formula = "=0"
        set_cell(ws, 39, 3 + c_idx, formula, font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    # Row 40: IRR — using Excel's IRR() on investor cash flows C39:M39
    set_cell(ws, 40, 1, "Investor IRR  (=IRR of cash flows)", font=BLACK, alignment=LEFT)
    set_cell(ws, 40, 3, "=IRR(C39:M39)", font=BLACK, alignment=CENTER, number_format=PCT_FORMAT)
    add_comment(ws, 40, 3,
        "IRR uses Excel's built-in IRR() function on the investor cash flow series.\n"
        "Cash flows: Year 0 = -Investment; Years 1-4 = 0; Year 5 = Exit Proceeds; Years 6-10 = 0.")

    # Row 41: Cash-on-Cash multiple
    set_cell(ws, 41, 1, "Cash-on-Cash Multiple  (Exit Proceeds ÷ Investment)", font=BLACK, alignment=LEFT)
    set_cell(ws, 41, 3, "=$C$38/$C$34", font=BLACK, alignment=RIGHT, number_format='0.00"x"')

    # ── Section E: Summary Metrics ─────────────────────────────────────────────
    write_section_header(ws, 43, "E. SUMMARY METRICS", num_cols=14)

    summary_rows = [
        (44, "Equity Value Today (£)",       f"={exit_col_letter}31",  GBP_FORMAT),
        (45, "Investor IRR",                 "=$C$40",                 PCT_FORMAT),
        (46, "Cash-on-Cash Multiple",        "=$C$41",                 '0.00"x"'),
        (47, "Exit Proceeds (Year 5, £)",    "=$C$38",                 GBP_FORMAT),
    ]
    for row_num, label, formula, fmt in summary_rows:
        set_cell(ws, row_num, 1, label, font=BLACK, alignment=LEFT)
        set_cell(ws, row_num, 3, formula, font=BLACK, alignment=RIGHT, number_format=fmt,
                 fill=TOTAL_FILL)

    ws.freeze_panes = "C4"
    return ws


# ==============================================================================
# 7. CREATE CAP TABLE SHEET
# ==============================================================================

def create_cap_table_sheet(wb: openpyxl.Workbook, assumptions: dict):
    """
    Builds the Cap Table tab — tracks equity ownership across all funding rounds.

    Funding rounds: Pre-seed T1, Pre-seed T2, Seed, Series A T1, Series A T2, Series B

    Shares issued formula (correct cumulative dilution):
        new_shares = cumulative_shares_before_round / (1 - ownership_pct) - cumulative_shares_before_round

    Row map:
      3   — Section A: INITIAL CAP TABLE header
      4   — Column headers
      5   — Founders row (1,000,000 shares, 100%)
      7   — Section B: FUNDING ROUNDS header
      8   — Column headers
      9-14 — Six funding round rows
      16  — Section C: CUMULATIVE CAP TABLE header
      17  — Column headers (shareholders × rounds)
      18-24 — Shareholder rows (Founders + 6 investors)
      25  — Total shares row
      27  — Section D: FINAL OWNERSHIP % header
      28  — Column headers
      29-35 — Final ownership rows
      36  — Total / check row

    Assumptions sheet column mapping for equity raises (row 52):
      C52 = Pre-seed T1 (£500K, 2025-10 raise)
      D52 = Pre-seed T2 + Seed (both map to the 2027 model raise slot, £1.5M)
      E52 = Series A T1 (£4M, 2028)
      F52 = Series A T2 (£0, 2029)
      G52 = Series B (£9M, 2030)
    """
    ws = wb.create_sheet("Cap Table")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    for c_idx in range(4, 12):
        ws.column_dimensions[col(c_idx)].width = 15

    rounds = assumptions["funding_rounds"]   # list of 6 dicts
    founder_shares = assumptions["founder_shares"]
    fx_rate = assumptions["fx_rate_usd_gbp"]

    # ── Title ──────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 24
    set_cell(ws, 1, 1, "CAPITALISATION TABLE  (CAP TABLE)", font=TITLE_FONT, fill=SECTION_FILL)
    for c in range(2, 12):
        ws.cell(row=1, column=c).fill = SECTION_FILL

    set_cell(ws, 2, 1, "Share counts are fractional — rounded for display only", font=BLACK)
    # FX rate input (top-right area)
    set_cell(ws, 2, 9, "GBP/USD FX Rate:", font=BLACK, alignment=RIGHT)
    set_cell(ws, 2, 10, fx_rate, font=BLUE, alignment=CENTER, number_format="0.0000")
    add_comment(ws, 2, 10,
        "GBP/USD FX rate used for $ conversions.\n"
        "Hardcoded fallback = 0.7577. Update cell directly to refresh $ columns.")

    # ── Section A: Initial Cap Table ──────────────────────────────────────────
    write_section_header(ws, 3, "A. INITIAL CAP TABLE  (Pre-Funding)", num_cols=11)

    hdr_a = ["Shareholder", "Shares", "Ownership %"]
    for i, h in enumerate(hdr_a, start=1):
        set_cell(ws, 4, i, h, font=BOLD_BLACK, alignment=CENTER, border=THIN_BORDER)

    set_cell(ws, 5, 1, "Founders",       font=BLUE, alignment=LEFT)
    set_cell(ws, 5, 2, founder_shares,   font=BLUE, alignment=RIGHT, number_format=NUM_FORMAT)
    set_cell(ws, 5, 3, "=B5/B5",         font=BLACK, alignment=CENTER, number_format=PCT_FORMAT)
    add_comment(ws, 5, 2, f"Founder shares: {founder_shares:,} (from Assumptions)")

    # ── Section B: Funding Rounds ──────────────────────────────────────────────
    write_section_header(ws, 7, "B. FUNDING ROUNDS", num_cols=11)

    round_hdrs = [
        "Round", "Date", "Ownership %", "Shares Issued",
        "Amount (£)", "Amount ($)", "Valuation (£)", "Valuation ($)"
    ]
    for i, h in enumerate(round_hdrs, start=1):
        set_cell(ws, 8, i, h, font=BOLD_BLACK, alignment=CENTER, border=THIN_BORDER)

    # Equity raises in Assumptions row 52, columns C-G
    # Mapping: round index → Assumptions column (per problem spec)
    # Pre-seed T1 (0)→C52 (£500K, 2025-10 raise), Pre-seed T2 (1)→D52 (£1.5M, same D column
    # as Seed because both fall in the model's 2027 raise slot), Seed (2)→D52,
    # Series A T1 (3)→E52 (£4M), Series A T2 (4)→F52 (£0), Series B (5)→G52 (£9M)
    equity_raise_cols = ["C", "D", "D", "E", "F", "G"]

    add_comment(ws, 8, 4,
        "Shares Issued formula (correct cumulative dilution):\n"
        "  new_shares = cumulative_shares / (1 - ownership_pct) - cumulative_shares\n"
        "  Uses running total of ALL shares outstanding before this round.\n"
        "  This ensures investors receive exactly their target post-round ownership %.")

    # Compute share data rows (9-14)
    # We need running totals — build these as Excel formulas referencing each other
    # Row 9 = round 0 (Pre-seed T1), row 10 = round 1 (Pre-seed T2), ...
    for r_idx, rnd in enumerate(rounds):
        data_row = 9 + r_idx   # rows 9..14

        set_cell(ws, data_row, 1, rnd["name"], font=BLUE, alignment=LEFT)
        set_cell(ws, data_row, 2, rnd["date"], font=BLUE, alignment=CENTER)

        # Ownership %
        set_cell(ws, data_row, 3, rnd["ownership_pct"],
                 font=BLUE, alignment=CENTER, number_format="0.0000%")

        # Shares issued: cumulative_before / (1 - pct) - cumulative_before
        # Cumulative shares before this round:
        #   round 0: B5 (founder shares)
        #   round n: B5 + sum of D9:D(data_row-1)
        if r_idx == 0:
            cum_before = "B5"
        else:
            cum_before = f"B5+SUM(D9:D{data_row - 1})"

        shares_formula = f"=({cum_before})/(1-C{data_row})-({cum_before})"
        set_cell(ws, data_row, 4, shares_formula, font=BLACK, alignment=RIGHT, number_format=NUM_FORMAT)

        # Amount raised (£) — linked to Assumptions!<col>52
        assump_col = equity_raise_cols[r_idx]
        amount_formula = f"=Assumptions!{assump_col}52"
        set_cell(ws, data_row, 5, amount_formula, font=GREEN, alignment=RIGHT, number_format=GBP_FORMAT)

        # Amount raised ($) = Amount(£) / fx_rate
        set_cell(ws, data_row, 6, f"=E{data_row}/$J$2",
                 font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

        # Valuation (£) = Amount(£) / ownership_pct
        set_cell(ws, data_row, 7, f"=E{data_row}/C{data_row}",
                 font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

        # Valuation ($)
        set_cell(ws, data_row, 8, f"=G{data_row}/$J$2",
                 font=BLACK, alignment=RIGHT, number_format=GBP_FORMAT)

    # FX rate reference cell (J2 used above)
    set_cell(ws, 2, 10, fx_rate, font=BLUE, alignment=CENTER, number_format="0.0000")

    # ── Section C: Cumulative Cap Table Matrix ─────────────────────────────────
    write_section_header(ws, 16, "C. CUMULATIVE CAP TABLE  (Shares after each round)", num_cols=11)

    # Column headers: Shareholder | Founders | Pre-seed T1 | ... | Series B
    set_cell(ws, 17, 1, "Shareholder", font=BOLD_BLACK, alignment=CENTER, border=THIN_BORDER)
    set_cell(ws, 17, 2, "Founders",    font=BOLD_BLACK, alignment=CENTER, border=THIN_BORDER)
    for r_idx, rnd in enumerate(rounds):
        set_cell(ws, 17, 3 + r_idx, rnd["name"],
                 font=BOLD_BLACK, alignment=CENTER, border=THIN_BORDER)

    # Shareholder rows 18-24: Founders (18), then each investor round (19-24)
    # For each column = cumulative shares held by that shareholder after that column's round

    # Row 18: Founders — always hold their original B5 shares
    set_cell(ws, 18, 1, "Founders", font=BLACK, alignment=LEFT)
    # After each round founders hold same shares (no dilution in share count, just %)
    for col_idx in range(7):   # 7 columns: Founders, T1..Series B
        set_cell(ws, 18, 2 + col_idx, "=$B$5", font=BLACK, alignment=RIGHT, number_format=NUM_FORMAT)

    # Rows 19-24: Each investor (holds 0 before their round, shares_issued after)
    for r_idx, rnd in enumerate(rounds):
        investor_row = 19 + r_idx   # 19..24
        set_cell(ws, investor_row, 1, rnd["name"], font=BLACK, alignment=LEFT)

        for col_idx in range(7):   # columns 2-8 = Founders round through Series B
            # col_idx 0 = "Founders only" column, col_idx 1 = after T1, etc.
            if col_idx < r_idx + 1:
                # Before this investor's round — they hold 0
                set_cell(ws, investor_row, 2 + col_idx, 0, font=BLACK, alignment=RIGHT, number_format=NUM_FORMAT)
            else:
                # After their round — hold the shares they were issued (D9+r_idx row)
                shares_row = 9 + r_idx
                set_cell(ws, investor_row, 2 + col_idx, f"=$D${shares_row}",
                         font=BLACK, alignment=RIGHT, number_format=NUM_FORMAT)

    # Row 25: Column totals
    write_subheader(ws, 25, "TOTAL SHARES", num_cols=8)
    for col_idx in range(7):
        col_letter = col(2 + col_idx)
        formula = f"=SUM({col_letter}18:{col_letter}24)"
        set_cell(ws, 25, 2 + col_idx, formula,
                 font=BOLD_BLACK, fill=SUBSECT_FILL, alignment=RIGHT, number_format=NUM_FORMAT)

    # ── Section D: Final Ownership % ──────────────────────────────────────────
    write_section_header(ws, 27, "D. FINAL OWNERSHIP %  (After All Rounds)", num_cols=11)

    set_cell(ws, 28, 1, "Shareholder",  font=BOLD_BLACK, alignment=CENTER, border=THIN_BORDER)
    set_cell(ws, 28, 2, "Shares Held",  font=BOLD_BLACK, alignment=CENTER, border=THIN_BORDER)
    set_cell(ws, 28, 3, "Ownership %",  font=BOLD_BLACK, alignment=CENTER, border=THIN_BORDER)

    # Total shares after all rounds is in column H row 25 (col_idx=6 → col 8=H)
    total_col_letter = col(2 + 6)   # H

    shareholder_rows = [
        (29, "Founders",     "=$B$5"),
    ]
    for r_idx, rnd in enumerate(rounds):
        shares_row = 9 + r_idx
        shareholder_rows.append((30 + r_idx, rnd["name"], f"=$D${shares_row}"))

    for row_num, name, shares_formula in shareholder_rows:
        set_cell(ws, row_num, 1, name, font=BLACK, alignment=LEFT)
        set_cell(ws, row_num, 2, shares_formula, font=BLACK, alignment=RIGHT, number_format=NUM_FORMAT)
        # Ownership % = shares / total_shares
        set_cell(ws, row_num, 3, f"=B{row_num}/{total_col_letter}25",
                 font=BLACK, alignment=CENTER, number_format=PCT_FORMAT)

    # Total check row
    write_subheader(ws, 36, "TOTAL", num_cols=4)
    set_cell(ws, 36, 2, f"={total_col_letter}25",
             font=BOLD_BLACK, fill=SUBSECT_FILL, alignment=RIGHT, number_format=NUM_FORMAT)
    set_cell(ws, 36, 3, "=SUM(C29:C35)",
             font=BOLD_BLACK, fill=SUBSECT_FILL, alignment=CENTER, number_format=PCT_FORMAT)
    add_comment(ws, 36, 3, "Should equal 100%. If not, check shares-issued formulas.")

    ws.freeze_panes = "B4"
    return ws


# ==============================================================================
# 8. APPLY GLOBAL FORMATTING
# ==============================================================================

def apply_global_formatting(wb: openpyxl.Workbook):
    """
    Applies consistent global formatting across all worksheets:
    - Calibri 10pt font throughout
    - Row heights for readability
    - Freeze panes (already set per sheet)
    - Tab colours for sheet identification
    - Print settings
    """
    tab_colours = {
        "Assumptions":      "1F4E79",   # dark blue
        "Scenarios":        "2E75B6",   # medium blue
        "Income Statement": "375623",   # dark green
        "Balance Sheet":    "843C0C",   # dark orange/brown
        "Cash Flow":        "7030A0",   # purple
        "Valuation":        "C65911",   # burnt orange
        "Cap Table":        "7030A0",   # purple
    }

    for sheet_name, hex_colour in tab_colours.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.sheet_properties.tabColor = hex_colour
            ws.print_title_rows = "1:3"
            ws.page_setup.fitToPage  = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.orientation = "landscape"

    # Ensure "Assumptions" is first tab, then Scenarios, then statements
    desired_order = ["Assumptions", "Scenarios", "Income Statement", "Balance Sheet", "Cash Flow", "Valuation", "Cap Table"]
    for i, name in enumerate(desired_order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=wb.sheetnames.index(name) - i)


# ==============================================================================
# 7. MAIN
# ==============================================================================

def main():
    """Orchestrates the model generation and saves the workbook."""
    wb = openpyxl.Workbook()

    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    print("Building Joule Financial Model...")

    print("  [1/8] Creating Assumptions sheet...")
    create_assumptions_sheet(wb, ASSUMPTIONS)

    print("  [2/8] Creating Scenarios sheet...")
    create_scenarios_sheet(wb)

    print("  [3/8] Building Income Statement...")
    build_income_statement(wb)

    print("  [4/8] Building Balance Sheet...")
    build_balance_sheet(wb)

    print("  [5/8] Building Cash Flow Statement...")
    build_cash_flow(wb)

    print("  [6/8] Building Valuation sheet...")
    create_valuation_sheet(wb, ASSUMPTIONS)

    print("  [7/8] Building Cap Table...")
    create_cap_table_sheet(wb, ASSUMPTIONS)

    print("  [8/8] Applying global formatting...")
    apply_global_formatting(wb)

    output_path = "financial_model.xlsx"
    wb.save(output_path)
    print(f"\n✓ Model saved to: {output_path}")
    print("\nKey notes:")
    print("  • Change scenario via Assumptions!C4 dropdown (Worst / Base / Best)")
    print("  • All financial assumptions can be updated in the 'Assumptions' tab")
    print("  • Balance Sheet check row (row 22) should show £0 in all columns")
    print("  • kWh revenue starts in 2029 (3-year project build lag)")
    print("  • Tax losses carried forward with £5M annual deduction limit")


if __name__ == "__main__":
    main()
